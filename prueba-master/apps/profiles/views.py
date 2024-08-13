from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.shortcuts import redirect, render
from django.db.models import Q, Sum
import os
from apps.profiles.resources import *
from apps.profiles.form import *

# Libreria exportaciones
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import csv
# Libreria importacion
from tablib import Dataset
import sys


def user_validate_group(roles=[]):
	def decorator(view_func):
		def arguments(request, *args, **kwargs):
			if request.user.groups.exists():
				if request.user.groups.all()[0].name in roles:
					return view_func(request, *args, **kwargs)
			return redirect('profiles:login')
		return arguments
	return decorator


def login_user(request):
    context={}
    if request.POST:
        form=AuthenticationForm(None, request.POST)
        username = request.POST['username']
        password = request.POST['password']
        if form.is_valid():
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('profiles:home')
            else:
                form.add_error(None, "Usuario no encontrado")
        context['form']=form
    return render(request, "general/login.html", context)

def logout_user(request):
    logout(request)
    return redirect('/profiles/login')

@login_required(login_url='/profiles/login')
# @user_validate_group(roles=['Administrator', 'Authorizer'])
def home_view(request):
    return render(request, "general/home.html")


@login_required(login_url='/profiles/login')
def read_academic_space_view(request):
    academic_spaces = AcademicSpace.objects.order_by('name')
    context = {}       
    if request.method == 'GET':
        if request.GET.get('search') != None:
            academic_spaces = academic_spaces.filter(Q(name__icontains=request.GET.get('search')))
           
        if request.GET.get('action') == 'download':
            dataset = AcademicSpaceExport().export(academic_spaces)
            response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename="AcademicSpace.xls"'
            return response
    if request.method == 'POST':
        if request.user.groups.all()[0].id == 1:
            if request.POST.get('action') == 'upload':
                try:
                    academic_import = AcademicSpaceImport()
                    dataset = Dataset()
                    dataset.load(request.FILES['upload_file'].read(),format='xls')
                    result = academic_import.import_data(dataset, dry_run=True, raise_errors=False, collect_failed_rows = False)
                    if not result.has_errors():
                        academic_import.import_data(dataset, dry_run=False)
                    context['result'] = result
                except Exception as e:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]  

    page = request.GET.get('page', 1)
    paginator = Paginator(academic_spaces,30)
    try:
        academic_spaces = paginator.page(page)
    except PageNotAnInteger:
        academic_spaces = paginator.page(1)
    except EmptyPage:
        academic_spaces = paginator.page(paginator.num_pages)
    context['academic_spaces']=academic_spaces
    context['amount_total']=paginator.count
    return render(request, "profiles/readAcademicSpace.html", context)



@login_required(login_url='/profiles/login')
def read_repetition_view(request):
    repetitions = Repetition.objects.order_by('id')
    academic_spaces = AcademicSpace.objects.all()
    context = {}       
    if request.method == 'GET':
        if request.GET.get('search') != None:
            academic_spaces = academic_spaces.filter(Q(name__icontains=request.GET.get('search')))
    if request.method == 'POST':
        if request.user.groups.all()[0].id == 1:
            if request.POST.get('action') == 'upload':
                try:
                    repetition_import = RepetitionImport()
                    dataset = Dataset()
                    dataset.load(request.FILES['upload_file'].read(),format='xls')
                    result = repetition_import.import_data(dataset, dry_run=True, raise_errors=False, collect_failed_rows = False)
                    if not result.has_errors():
                        repetition_import.import_data(dataset, dry_run=False)
                        print('no hubo error',result)
                    else:
                        print('si error',result.has_errors())
                        print('error',result)
                    context['result'] = result
                except Exception as e:
                    exc_type, exc_obj, exc_tb = sys.exc_info()
                    fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                    print(exc_type, exc_obj, exc_tb)
                    print(fname)
    data = []
    max_repeats = 0

    for space in academic_spaces:
        repetitions = Repetition.objects.filter(academic_space=space).order_by('-times_repeated')
        total_repitentes = repetitions.aggregate(total=Sum('student_count'))['total'] or 0

        # Calcular el máximo número de repeticiones para ajustar las cabeceras
        max_repeats = max(max_repeats, repetitions.count())

        # Construir la fila
        row = [space.name]
        for repetition in repetitions:
            row.extend([repetition.times_repeated, repetition.student_count])
        row.append(total_repitentes)
        data.append(row)

    # Crear las cabeceras dinámicas
    headers = ["ESPACIO ACADÉMICO"]
    for i in range(max_repeats):
        headers.extend([
            "NUMERO DE VECES QUE HA REPETIDO UN ESTUDIANTE", 
            "NUMERO DE ESTUDIANTES QUE HAN REPETIDO ESTAS VECES"
        ])
    headers.append("TOTAL REPITENTES")
    graph_data=data
    if request.GET.get('action') == 'download_xls':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Repeticiones"

            # Escribir las cabeceras
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws['{}{}'.format(col_letter, 1)] = header

            # Escribir los datos de la tabla
            for row_num, row in enumerate(data, 2):
                for col_num, cell_value in enumerate(row, 1):
                    col_letter = get_column_letter(col_num)
                    ws['{}{}'.format(col_letter, row_num)] = cell_value

            # Configurar la respuesta HTTP para descargar el archivo
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=repeticiones.xlsx'
            wb.save(response)
            return response
            # dataset = RepetitionExport().export(repetitions)
            # response = HttpResponse(dataset.xls, content_type='application/vnd.ms-excel')
            # response['Content-Disposition'] = 'attachment; filename="Repetition.xls"'
    if request.GET.get('action') == 'download_txt':
        # Configurar la respuesta HTTP para descargar el archivo
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=repeticiones.txt'

        writer = csv.writer(response, delimiter='\t')  # Usar tabulaciones como delimitador para formato .txt

        # Escribir las cabeceras
        writer.writerow(headers)

        # Escribir los datos de la tabla
        for row in data:
            writer.writerow(row)

        return response
        # dataset = RepetitionExport().export(repetitions)
        # response = HttpResponse(dataset.csv, content_type='text/plain')
        # response['Content-Disposition'] = 'attachment; filename="Estudiantes.txt"'
    page = request.GET.get('page', 1)
    paginator = Paginator(data,20)
    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)
    context['data']=data
    context['graph_data']=graph_data
    context['headers']=headers
    context['amount_total']=paginator.count
    return render(request, "profiles/readRepetition.html", context)

